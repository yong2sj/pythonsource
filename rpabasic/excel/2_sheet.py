from openpyxl import Workbook

# workbook 객체 생성
wb = Workbook()

# 새로운 시트 생성
ws1 = wb.create_sheet()
ws1.title = "MySheet"
ws1.sheet_properties.tabColor = "ff66ff"

ws2 = wb.create_sheet("급여명세", 2)

# 생성된 시트의 모든 이름 출력
print(wb.sheetnames)

# 기존 시트에 접근하기
new_ws = wb["MySheet"]
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

wb.save("./rpabasic/excel/sample.xlsx")

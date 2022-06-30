# 함수 삽입
from openpyxl import Workbook
from datetime import datetime

wb = Workbook()
ws = wb.active

ws["a1"] = datetime.today()
ws["a2"] = "=sum(1,2,3)"
ws["a3"] = "=average(1,2,3)"

ws["a4"] = 10
ws["a5"] = 20
ws["a6"] = "=sum(a4:a5)"

wb.save("./rpabasic/excel/formula.xlsx")

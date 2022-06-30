from openpyxl import load_workbook

wb = load_workbook("./rpabasic/excel/range.xlsx")
ws = wb.active

# for row in ws.iter_rows(min_row=2): # 제목행이 필요없다면 2행부터 출력
#     print(row[0].value, row[1].value, row[2].value)

# 영어점수가 80이상만 출력
# for row in ws.iter_rows(min_row=2):
#     if row[1].value >= 80:
#         print(row[0].value, row[1].value, row[2].value)

# 셀 안에 영어라는 문자가 입력된 셀이 있는 경우 컴퓨터 변경
for row in ws.iter_rows(min_row=1):
    for cell in row:
        if cell.value == "영어":
            cell.value = "컴퓨터"

wb.save("./rpabasic/excel/range.xlsx")

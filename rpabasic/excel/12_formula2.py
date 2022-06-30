# 함수 삽입
from openpyxl import load_workbook
from datetime import datetime

# 수식이 들어있는 경우 수식 그대로 가져옴
# wb = load_workbook("./rpabasic/excel/formula.xlsx")

wb = load_workbook("./rpabasic/excel/formula.xlsx", data_only=True)
ws = wb.active

for row in ws.values:
    for cell in row:
        print(cell)

# wb.save("./rpabasic/excel/formula.xlsx")

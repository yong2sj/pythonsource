# 셀 스타일
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, PatternFill, Alignment
from openpyxl.styles.borders import Side

wb = load_workbook("./rpabasic/excel/range.xlsx")
ws = wb.active

# 셀 너비 조절
ws.column_dimensions["A"].width = 5

# 행 높이 조절
ws.row_dimensions[1].height = 50

# 폰트 스타일
a1 = ws["A1"]
c1 = ws["C1"]
d1 = ws["D1"]

a1.font = Font(color="ff0000", italic=True, bold=True)
c1.font = Font(color="cc33ff", name="Arial", strike=True)
d1.font = Font(color="0000ff", size=20, underline="single")

# 테두리 적용
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

a1.border = thin_border
c1.border = thin_border
d1.border = thin_border

# 셀 배경색
for row in ws.rows:
    for cell in row:
        # 셀 정렬
        cell.alignment = Alignment(horizontal="center", vertical="center")

        if cell.column == 1:
            continue
        if isinstance(cell.value, int) and cell.value > 80:
            cell.fill = PatternFill(fgColor="00ff00", fill_type="solid")
            cell.font = Font(color="ff0000")

# 틀고정
ws.freeze_panes = "C2"

wb.save("./rpabasic/excel/range.xlsx")

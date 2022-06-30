# 셀병합
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

# 병합 셀 범위 지정
ws.merge_cells("B2:D2")
ws["B2"].value = "Merged cell"

wb.save("./rpabasic/excel/merge.xlsx")

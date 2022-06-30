# 셀병합
from openpyxl import load_workbook

wb = load_workbook("./rpabasic/excel/merge.xlsx")
ws = wb.active

# 병합 셀 해제
ws.unmerge_cells("B2:D2")

wb.save("./rpabasic/excel/merge.xlsx")

# gmarket best - 컴퓨터/전자 항목 추출 + 엑셀저장
from ctypes import alignment
from turtle import color
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime  # 오늘날짜

# 엑셀
wb = Workbook()
ws = wb.active
ws.title = "컴퓨터전자100"

# 상품명 너비 조절
ws.column_dimensions["B"].width = 50
# 가격
ws.column_dimensions["C"].width = 15
# 회사명
ws.column_dimensions["D"].width = 20
# 상세주소
ws.column_dimensions["E"].width = 80
# 제목 행 추가
ws.append(["번호", "상품명", "가격", "회사명", "상세정보url"])


url = "http://corners.gmarket.co.kr/Bestsellers?viewType=G&groupCode=G06"

res = requests.get(url)
soup = BeautifulSoup(res.text, "lxml")

# 1~100위 아이템 찾기
best_list = soup.select("div.best-list li")

for idx, item in enumerate(best_list):
    # 상품명
    title = item.select_one("a.itemname")
    # 가격
    price = item.select_one("div.s-price span").get_text()

    # 순위, 상품명, 가격, 상품상세정보url
    # print(idx + 1, title.get_text(), price, title["href"])

    # title["href"] 를 이용해서 requests.get() 요청
    product_url = title["href"]
    res = requests.get(product_url)
    detail_soup = BeautifulSoup(res.text, "lxml")

    # 회사명 추출(회사명 없으면 셀러명 추출)
    company = detail_soup.select_one("span.text__brand > span.text")

    # 제조사가 없는 경우 셀러명 추출
    if not company:
        company = detail_soup.select_one("span.text__seller > a")

    # 순위, 상품명, 가격, 회사명, 상품상세정보url
    # print(idx + 1, title.get_text(), price, company.get_text(), product_url)

    ws.append([idx + 1, title.get_text(), price, company.get_text(), product_url])
    # 하이퍼링크 걸기(제목행 제외)
    ws.cell(row=idx + 1, column=5).hyperlink = product_url

# 파일명 gmaket_오늘날짜.xlsx
today = datetime.now().strftime("%y%m%d")
filename = f"gmarket_{today}.xlsx"

# 제목 행 서식 지정
font = Font(name="Tahoma", size=14, color="01579b")
alignment = Alignment(horizontal="center")

cell_a1 = ws["A1"]
cell_a1.alignment = alignment
cell_a1.font = font

cell_b1 = ws["B1"]
cell_b1.alignment = alignment
cell_b1.font = font

cell_c1 = ws["C1"]
cell_c1.alignment = alignment
cell_c1.font = font

cell_d1 = ws["D1"]
cell_d1.alignment = alignment
cell_d1.font = font

cell_e1 = ws["E1"]
cell_e1.alignment = alignment
cell_e1.font = font


# 엑셀 저장
wb.save("./rpabasic/crawl/download/" + filename)
